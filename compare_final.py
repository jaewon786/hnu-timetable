#!/usr/bin/env python3
"""Final comparison - correct sheet priority, normalized formats."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import re, os
import openpyxl

EXCEL_PATH = r'26-1 수강편람 (4차).xlsx'
COURSES_DIR = r'src\data\courses'
DAYS = '월화수목금토'

def normalize_id(raw_id):
    parts = raw_id.split('-')
    if len(parts) == 2:
        code, section = parts
        return f"{code}-{section.lstrip('0') or '0'}"
    return raw_id

def normalize_time(t):
    if not t: return ''
    t = str(t).strip()
    t = t.replace('.', ',')
    t = re.sub(r',[' + DAYS + r']', lambda m: '/' + m.group(0)[1], t)
    t = re.sub(r' [' + DAYS + r']', lambda m: '/' + m.group(0)[1], t)
    t = re.sub(r'\n[' + DAYS + r']', lambda m: '/' + m.group(0)[1], t)
    return t

def normalize_room(r):
    if not r: return ''
    r = str(r).strip()
    r = r.replace('\n', '/')
    # Remove trailing -0 from each room segment
    parts = r.split('/')
    normalized = []
    for p in parts:
        p = p.strip()
        if p.endswith('-0'):
            p = p[:-2]
        # Remove leading zeros for comparison (060334 -> 60334)
        p = p.lstrip('0') or '0'
        normalized.append(p)
    return '/'.join(normalized)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Build Excel data with priority: 교필 > 교선 > 전공 > 코드쉐어 > 마이크로디그리
excel = {}

# 교필
ws = wb['교필']
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    cat = ws.cell(r, 2).value
    code = ws.cell(r, 3).value
    section = ws.cell(r, 4).value
    credit = ws.cell(r, 5).value
    prof = ws.cell(r, 9).value
    time_val = ws.cell(r, 10).value
    room = ws.cell(r, 11).value
    if not name or not code or name == '과목명': continue
    nid = f"{code}-{str(section).lstrip('0') or '0'}"
    if nid not in excel:
        excel[nid] = {
            'sheet': '교필', 'name': str(name).strip(),
            'category': str(cat).strip() if cat else '',
            'creditDetail': str(credit).strip() if credit else '',
            'professor': str(prof).strip() if prof else '',
            'timeRaw': str(time_val).strip() if time_val else '',
            'roomRaw': str(room).strip() if room else '',
        }

# 교선
ws = wb['교선']
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    cat = ws.cell(r, 2).value
    code = ws.cell(r, 3).value
    section = ws.cell(r, 4).value
    credit = ws.cell(r, 5).value
    prof = ws.cell(r, 6).value
    time_val = ws.cell(r, 7).value
    room = ws.cell(r, 8).value
    if not name or not code or name == '교과목명': continue
    nid = f"{code}-{str(section).lstrip('0') or '0'}"
    if nid not in excel:
        excel[nid] = {
            'sheet': '교선', 'name': str(name).strip(),
            'category': str(cat).strip() if cat else '',
            'creditDetail': str(credit).strip() if credit else '',
            'professor': str(prof).strip() if prof else '',
            'timeRaw': str(time_val).strip() if time_val else '',
            'roomRaw': str(room).strip() if room else '',
        }

# 전공
ws = wb['전공']
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 5).value
    section = ws.cell(r, 6).value
    name = ws.cell(r, 7).value
    cat = ws.cell(r, 8).value
    credit = ws.cell(r, 9).value
    prof = ws.cell(r, 10).value
    time_val = ws.cell(r, 11).value
    room = ws.cell(r, 12).value
    if not name or not code or name == '과목 명': continue
    nid = f"{code}-{str(section).lstrip('0') or '0'}"
    if nid not in excel:
        excel[nid] = {
            'sheet': '전공', 'name': str(name).strip(),
            'category': str(cat).strip() if cat else '',
            'creditDetail': str(credit).strip() if credit else '',
            'professor': str(prof).strip() if prof else '',
            'timeRaw': str(time_val).strip() if time_val else '',
            'roomRaw': str(room).strip() if room else '',
        }

# 코드쉐어
ws = wb['코드쉐어']
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 4).value
    section = ws.cell(r, 5).value
    dept = ws.cell(r, 1).value
    cat = ws.cell(r, 8).value
    prof = ws.cell(r, 9).value
    time_val = ws.cell(r, 10).value
    room = ws.cell(r, 11).value
    if not code or str(code).strip() == '학수번호': continue
    nid = f"{code}-{str(section).lstrip('0') or '0'}"
    if nid not in excel:
        excel[nid] = {
            'sheet': '코드쉐어', 'name': str(dept).strip() if dept else '',
            'category': str(cat).strip() if cat else '',
            'professor': str(prof).strip() if prof else '',
            'timeRaw': str(time_val).strip() if time_val else '',
            'roomRaw': str(room).strip() if room else '',
        }

# 마이크로디그리
ws = wb['마이크로디그리']
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 2).value
    section = ws.cell(r, 4).value
    name = ws.cell(r, 3).value
    cat = ws.cell(r, 7).value
    prof = ws.cell(r, 8).value
    time_val = ws.cell(r, 9).value
    room = ws.cell(r, 10).value
    if not code or str(code).strip() == '학수번호': continue
    nid = f"{code}-{str(section).lstrip('0') or '0'}"
    if nid not in excel:
        excel[nid] = {
            'sheet': '마이크로디그리', 'name': str(name).strip() if name else '',
            'category': str(cat).strip() if cat else '',
            'professor': str(prof).strip() if prof else '',
            'timeRaw': str(time_val).strip() if time_val else '',
            'roomRaw': str(room).strip() if room else '',
        }

print(f"Excel unique IDs: {len(excel)}")

# Parse TS files
ts_files = ['core.ts', 'electives.ts', 'major_required.ts', 'major_elective.ts',
            'semester.ts', 'normal_electives.ts', 'teaching.ts', 'online.ts']
all_ts = {}
for f in ts_files:
    path = os.path.join(COURSES_DIR, f)
    with open(path, 'r', encoding='utf-8') as fh:
        content = fh.read()
    for m in re.finditer(r"id:\s*'([^']+)'", content):
        raw_id = m.group(1)
        nid = normalize_id(raw_id)
        start = content.rfind('{', 0, m.start())
        region = content[start:start+2000]
        name_m = re.search(r"name:\s*'([^']*?)'", region)
        prof_m = re.search(r"professors:\s*\[([^\]]*)\]", region)
        time_m = re.search(r"timeRaw:\s*'([^']*?)'", region)
        room_m = re.search(r"roomRaw:\s*'([^']*?)'", region)
        cat_m = re.search(r"category:\s*'([^']*?)'", region)
        credit_m = re.search(r"creditDetail:\s*'([^']*?)'", region)
        profs = re.findall(r"'([^']*?)'", prof_m.group(1)) if prof_m else []
        if nid not in all_ts:
            all_ts[nid] = {
                'file': f,
                'name': name_m.group(1) if name_m else '',
                'professors': profs,
                'timeRaw': time_m.group(1) if time_m else '',
                'roomRaw': room_m.group(1) if room_m else '',
                'category': cat_m.group(1) if cat_m else '',
                'creditDetail': credit_m.group(1) if credit_m else '',
            }

print(f"TS unique IDs: {len(all_ts)}")

excel_ids = set(excel.keys())
ts_ids = set(all_ts.keys())
missing = excel_ids - ts_ids
extra = ts_ids - excel_ids
common = excel_ids & ts_ids

print(f"Common: {len(common)}, Missing in TS: {len(missing)}, Extra in TS: {len(extra)}")

if missing:
    print(f"\n--- MISSING IN TS ---")
    for m in sorted(missing):
        e = excel[m]
        print(f"  {m}: {e['name']} [{e['sheet']}]")

if extra:
    print(f"\n--- EXTRA IN TS ---")
    for e in sorted(extra):
        t = all_ts[e]
        print(f"  {e}: {t['name']} [{t['file']}]")

# Field diffs
print(f"\n=== FIELD DIFFERENCES ===")
time_diffs = []
room_diffs = []
name_diffs = []
prof_diffs = []
cat_diffs = []

for cid in sorted(common):
    ec = excel[cid]
    tc = all_ts[cid]

    # Time (normalized)
    et = normalize_time(ec.get('timeRaw', ''))
    tt = tc.get('timeRaw', '')
    if et != tt:
        time_diffs.append((cid, ec.get('name',''), et, tt, tc['file'], ec['sheet']))

    # Room (normalized)
    er = normalize_room(ec.get('roomRaw', ''))
    tr = normalize_room(tc.get('roomRaw', ''))
    if er != tr:
        room_diffs.append((cid, ec.get('name',''), ec.get('roomRaw',''), tc.get('roomRaw',''), tc['file'], ec['sheet']))

    # Name (ignore I/II/III variants)
    en = ec.get('name', '')
    tn = tc.get('name', '')
    if en and tn and en != tn:
        en_n = en.replace('Ⅰ','I').replace('Ⅱ','II').replace('Ⅲ','III')
        tn_n = tn.replace('Ⅰ','I').replace('Ⅱ','II').replace('Ⅲ','III')
        if en_n != tn_n:
            name_diffs.append((cid, en, tn, tc['file'], ec['sheet']))

    # Professor (only check where Excel has a value)
    ep = ec.get('professor', '')
    tp = ','.join(tc.get('professors', []))
    if ep and ep != tp:
        prof_diffs.append((cid, ec.get('name',''), ep, tp, tc['file'], ec['sheet']))

    # Category (skip 기업가정신)
    ecat = ec.get('category', '')
    tcat = tc.get('category', '')
    if ecat and tcat and ecat != tcat and '기업가정신' not in ec.get('name', ''):
        cat_diffs.append((cid, ec.get('name',''), ecat, tcat, tc['file'], ec['sheet']))

print(f"\nTime diffs (after normalization): {len(time_diffs)}")
for cid, name, et, tt, f, s in time_diffs:
    print(f"  {cid} ({name}) [{s}->{f}]: Excel=\"{et}\" vs TS=\"{tt}\"")

print(f"\nRoom diffs (after normalization): {len(room_diffs)}")
for cid, name, er, tr, f, s in room_diffs[:30]:
    print(f"  {cid} ({name}) [{s}->{f}]: Excel=\"{er}\" vs TS=\"{tr}\"")
if len(room_diffs) > 30:
    print(f"  ... and {len(room_diffs)-30} more")

print(f"\nName diffs (excluding I/II/III): {len(name_diffs)}")
for cid, en, tn, f, s in name_diffs:
    print(f"  {cid} [{s}->{f}]: Excel=\"{en}\" vs TS=\"{tn}\"")

print(f"\nProfessor diffs (Excel non-empty): {len(prof_diffs)}")
for cid, name, ep, tp, f, s in prof_diffs[:30]:
    print(f"  {cid} ({name}) [{s}->{f}]: Excel=\"{ep}\" vs TS=\"{tp}\"")
if len(prof_diffs) > 30:
    print(f"  ... and {len(prof_diffs)-30} more")

print(f"\nCategory diffs: {len(cat_diffs)}")
for cid, name, ecat, tcat, f, s in cat_diffs:
    print(f"  {cid} ({name}) [{s}->{f}]: Excel=\"{ecat}\" vs TS=\"{tcat}\"")
