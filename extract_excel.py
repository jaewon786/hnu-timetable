import openpyxl
import json

wb = openpyxl.load_workbook(r"C:\Users\jaewo\Desktop\hnu-timetable\26-1 수강편람 (4차).xlsx", data_only=True)

result = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    result[sheet_name] = rows

with open("excel_data.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print("Done! Sheets:", list(result.keys()))
for k, v in result.items():
    print(f"  {k}: {len(v)} rows")
