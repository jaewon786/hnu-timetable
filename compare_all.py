import openpyxl
import json
import re
import os

wb = openpyxl.load_workbook(r"C:\Users\jaewo\Desktop\hnu-timetable\26-1 수강편람 (4차).xlsx", data_only=True)

# ============ 1. 교필 시트 파싱 ============
ws = wb["교필"]
header_row = 2  # Row 2 has headers
excel_core = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name:
        continue
    code = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ""
    section = str(ws.cell(r, 4).value).strip().zfill(2) if ws.cell(r, 4).value else ""
    category = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
    credit_detail = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else ""
    professor = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""
    time_raw = str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""
    room_raw = str(ws.cell(r, 11).value).strip() if ws.cell(r, 11).value else ""
    
    id_key = f"{code}-{section}"
    excel_core.append({
        "id": id_key, "code": code, "section": section,
        "name": str(name).strip(), "category": category,
        "creditDetail": credit_detail, "professor": professor,
        "timeRaw": time_raw, "roomRaw": room_raw
    })

# ============ 2. 교선 시트 파싱 ============
ws = wb["교선"]
excel_elective = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name:
        continue
    category = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
    code = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ""
    section = str(ws.cell(r, 4).value).strip().zfill(2) if ws.cell(r, 4).value else ""
    credit_detail = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else ""
    professor = str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ""
    time_raw = str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ""
    room_raw = str(ws.cell(r, 8).value).strip() if ws.cell(r, 8).value else ""
    note = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""
    
    id_key = f"{code}-{section}"
    excel_elective.append({
        "id": id_key, "code": code, "section": section,
        "name": str(name).strip(), "category": category,
        "creditDetail": credit_detail, "professor": professor,
        "timeRaw": time_raw, "roomRaw": room_raw, "note": note
    })

# ============ 3. 전공 시트 파싱 ============
ws = wb["전공"]
# Headers at row 2: 단과대학, 학부/학과, 전공, 학년, 학수번호, 분반, 과목 명, 이수구분, 학점, 담당교수, 강의시간, 강의실, 비고
excel_major = []
for r in range(3, ws.max_row + 1):
    college = ws.cell(r, 1).value
    if not college:
        continue
    dept = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
    major = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ""
    year = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else ""
    code = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else ""
    section = str(ws.cell(r, 6).value).strip().zfill(2) if ws.cell(r, 6).value else ""
    name = str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ""
    category = str(ws.cell(r, 8).value).strip() if ws.cell(r, 8).value else ""
    credit_detail = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""
    professor = str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""
    time_raw = str(ws.cell(r, 11).value).strip() if ws.cell(r, 11).value else ""
    room_raw = str(ws.cell(r, 12).value).strip() if ws.cell(r, 12).value else ""
    note = str(ws.cell(r, 13).value).strip() if ws.cell(r, 13).value else ""
    
    id_key = f"{code}-{section}"
    excel_major.append({
        "id": id_key, "code": code, "section": section,
        "name": name, "college": str(college).strip(),
        "department": dept, "major": major, "year": year,
        "category": category, "creditDetail": credit_detail,
        "professor": professor, "timeRaw": time_raw, "roomRaw": room_raw,
        "note": note
    })

# ============ 4. 코드쉐어 시트 파싱 ============
# 코드쉐어 헤더: 교과목명(1), 주관대학(2), 주관학과(3), 학수번호(4), 분반(5), 학강실(6), 학년(7), 이수구분(8), 담당교원(9), 강의시간(10), 강의실(11)
ws = wb["코드쉐어"]
excel_codeshare = []
for r in range(3, ws.max_row + 1):
    name = ws.cell(r, 1).value
    if not name:
        continue
    code = str(ws.cell(r, 4).value).strip() if ws.cell(r, 4).value else ""
    section = str(ws.cell(r, 5).value).strip().zfill(2) if ws.cell(r, 5).value else ""
    credit_detail = str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ""
    category = str(ws.cell(r, 8).value).strip() if ws.cell(r, 8).value else ""
    professor = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""
    time_raw = str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""
    room_raw = str(ws.cell(r, 11).value).strip() if ws.cell(r, 11).value else ""
    
    id_key = f"{code}-{section}"
    excel_codeshare.append({
        "id": id_key, "code": code, "section": section,
        "name": str(name).strip(), "category": category,
        "creditDetail": credit_detail, "professor": professor,
        "timeRaw": time_raw, "roomRaw": room_raw
    })

# ============ 5. 마이크로디그리 시트 파싱 ============
# 마이크로디그리 헤더: 마이크로디그리명(1), 학수번호(2), 교과목명(3), 분반(4), 학-강-실(5), 학년(6), 이수구분(7), 담당교원(8), 강의시간(9), 강의실(10), 비고(11)
ws = wb["마이크로디그리"]
excel_micro = []
for r in range(3, ws.max_row + 1):
    micro_name = ws.cell(r, 1).value
    if not micro_name:
        continue
    code = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ""
    name = str(ws.cell(r, 3).value).strip() if ws.cell(r, 3).value else ""
    section = str(ws.cell(r, 4).value).strip().zfill(2) if ws.cell(r, 4).value else ""
    credit_detail = str(ws.cell(r, 5).value).strip() if ws.cell(r, 5).value else ""
    year = str(ws.cell(r, 6).value).strip() if ws.cell(r, 6).value else ""
    category = str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else ""
    professor = str(ws.cell(r, 8).value).strip() if ws.cell(r, 8).value else ""
    time_raw = str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""
    room_raw = str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""
    note = str(ws.cell(r, 11).value).strip() if ws.cell(r, 11).value else ""
    
    id_key = f"{code}-{section}"
    excel_micro.append({
        "id": id_key, "code": code, "section": section,
        "name": name, "microdegreeName": str(micro_name).strip(),
        "category": category, "creditDetail": credit_detail,
        "professor": professor, "timeRaw": time_raw, "roomRaw": room_raw,
        "note": note
    })

# ============ 소스 파일에서 id 추출 ============
def extract_ids_from_ts(filepath):
    """TS 파일에서 id 필드 추출"""
    ids = {}
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    
    # Extract id, name, timeRaw, roomRaw, professor from TS objects
    # Pattern for id
    id_pattern = re.compile(r"id:\s*'([^']*)'")
    name_pattern = re.compile(r"name:\s*'([^']*)'")
    time_pattern = re.compile(r"timeRaw:\s*'([^']*)'")
    room_pattern = re.compile(r"roomRaw:\s*'([^']*)'")
    professor_pattern = re.compile(r"professors:\s*\[([^\]]*)\]")
    category_pattern = re.compile(r"category:\s*'([^']*)'")
    credit_pattern = re.compile(r"creditDetail:\s*'([^']*)'")
    
    # Split by object boundaries (approximate)
    objects = content.split("  {")
    result = []
    for obj in objects:
        id_m = id_pattern.search(obj)
        if not id_m:
            continue
        entry = {"id": id_m.group(1)}
        
        name_m = name_pattern.search(obj)
        if name_m:
            entry["name"] = name_m.group(1)
        
        time_m = time_pattern.search(obj)
        if time_m:
            entry["timeRaw"] = time_m.group(1)
        
        room_m = room_pattern.search(obj)
        if room_m:
            entry["roomRaw"] = room_m.group(1)
        
        prof_m = professor_pattern.search(obj)
        if prof_m:
            entry["professors"] = prof_m.group(1).strip()
        
        cat_m = category_pattern.search(obj)
        if cat_m:
            entry["category"] = cat_m.group(1)
            
        cred_m = credit_pattern.search(obj)
        if cred_m:
            entry["creditDetail"] = cred_m.group(1)
        
        result.append(entry)
    
    return result

base = r"C:\Users\jaewo\Desktop\hnu-timetable\src\data\courses"
src_core = extract_ids_from_ts(os.path.join(base, "core.ts"))
src_electives = extract_ids_from_ts(os.path.join(base, "electives.ts"))
src_major_required = extract_ids_from_ts(os.path.join(base, "major_required.ts"))
src_major_elective = extract_ids_from_ts(os.path.join(base, "major_elective.ts"))
src_semester = extract_ids_from_ts(os.path.join(base, "semester.ts"))
src_normal_electives = extract_ids_from_ts(os.path.join(base, "normal_electives.ts"))
src_teaching = extract_ids_from_ts(os.path.join(base, "teaching.ts"))
src_online = extract_ids_from_ts(os.path.join(base, "online.ts"))

# 출력을 파일로 리디렉션
import sys
sys.stdout = open("compare_result_final.txt", "w", encoding="utf-8")

# ============ 비교 ============
def normalize_id(id_str):
    """ID 정규화: '11967-1' -> '11967-01'"""
    parts = id_str.split('-')
    if len(parts) == 2:
        return f"{parts[0]}-{parts[1].zfill(2)}"
    return id_str

def normalize_time(t):
    """시간 표기 정규화: '/' -> ',' 통일"""
    if not t:
        return t
    return t.replace('/', ',').replace(' ', '')

def normalize_name(n):
    """이름 정규화: 로마숫자 I->Ⅰ, II->Ⅱ 등 통일"""
    if not n:
        return n
    # 로마 숫자 정규화 (fullwidth -> ASCII)
    n = n.replace('Ⅰ', 'I').replace('Ⅱ', 'II').replace('Ⅲ', 'III').replace('Ⅳ', 'IV')
    return n

def compare_sheets(sheet_name, excel_data, src_data, src_file):
    print(f"\n{'='*60}")
    print(f"[비교] {sheet_name} (엑셀: {len(excel_data)}개, 소스: {len(src_data)}개)")
    print(f"{'='*60}")
    
    excel_ids = {e["id"] for e in excel_data}
    src_ids = {normalize_id(s["id"]) for s in src_data}
    
    # 엑셀에만 있는 과목 (소스에 없음 = 누락)
    missing_in_src = excel_ids - src_ids
    # 소스에만 있는 과목 (엑셀에 없음 = 삭제 필요)
    extra_in_src = src_ids - excel_ids
    
    if missing_in_src:
        print(f"\n[누락] 엑셀에 있지만 소스에 없는 과목 ({len(missing_in_src)}개):")
        for mid in sorted(missing_in_src):
            excel_entry = next((e for e in excel_data if e["id"] == mid), None)
            if excel_entry:
                print(f"   - {mid}: {excel_entry['name']} ({excel_entry.get('professor', '')})")
    
    if extra_in_src:
        print(f"\n[추가] 소스에 있지만 엑셀에 없는 과목 ({len(extra_in_src)}개):")
        for eid in sorted(extra_in_src):
            src_entry = next((s for s in src_data if normalize_id(s["id"]) == eid), None)
            if src_entry:
                print(f"   - {eid}: {src_entry.get('name', '?')}")
    
    # 공통 과목 중 데이터 차이 확인
    common_ids = excel_ids & src_ids
    diffs = []
    for cid in sorted(common_ids):
        excel_entry = next((e for e in excel_data if e["id"] == cid), None)
        src_entry = next((s for s in src_data if normalize_id(s["id"]) == cid), None)
        if not excel_entry or not src_entry:
            continue
        
        changes = []
        # 이름 비교 (로마숫자 정규화 후)
        src_name = normalize_name(src_entry.get("name", ""))
        excel_name = normalize_name(excel_entry.get("name", ""))
        if excel_name != src_name:
            changes.append(f"name: '{src_entry.get('name','')}'->'{excel_entry.get('name','')}'")
        # 시간 비교 (구분자 정규화 후)
        src_time = normalize_time(src_entry.get("timeRaw", ""))
        excel_time = normalize_time(excel_entry.get("timeRaw", ""))
        if excel_time != src_time:
            changes.append(f"timeRaw: '{src_entry.get('timeRaw','')}'->'{excel_entry.get('timeRaw','')}'")
        # 강의실 비교
        if excel_entry.get("roomRaw", "") != src_entry.get("roomRaw", ""):
            changes.append(f"roomRaw: '{src_entry.get('roomRaw','')}'->'{excel_entry.get('roomRaw','')}'")
        # 교수 비교
        excel_prof = excel_entry.get("professor", "")
        src_prof = src_entry.get("professors", "").replace("'", "").replace('"', '').replace(' ', '').strip()
        excel_prof_norm = excel_prof.replace(' ', '')
        if excel_prof_norm and src_prof and excel_prof_norm != src_prof:
            changes.append(f"professor: '{src_entry.get('professors','')}'->'{excel_prof}'")
        # 학점 비교
        if excel_entry.get("creditDetail", "") != src_entry.get("creditDetail", ""):
            changes.append(f"creditDetail: '{src_entry.get('creditDetail','')}'->'{excel_entry.get('creditDetail','')}'")
        
        if changes:
            diffs.append((cid, changes))
    
    if diffs:
        print(f"\n[차이] 데이터가 다른 과목 ({len(diffs)}개):")
        for cid, changes in diffs[:30]:  # 최대 30개만
            print(f"   - {cid}: {', '.join(changes)}")
        if len(diffs) > 30:
            print(f"   ... 외 {len(diffs)-30}개")
    
    if not missing_in_src and not extra_in_src and not diffs:
        print("   [OK] 완벽히 일치합니다!")
    
    return missing_in_src, extra_in_src, diffs

# 교필 비교
compare_sheets("교필 (core.ts)", excel_core, src_core, "core.ts")

# 교선 비교 (교직 과목 제외 - teaching.ts에서 별도 관리)
excel_elective_no_teach = [e for e in excel_elective if "교직" not in e.get("category", "")]
excel_elective_teach = [e for e in excel_elective if "교직" in e.get("category", "")]
compare_sheets("교선-일반 (electives.ts)", excel_elective_no_teach, src_electives, "electives.ts")
compare_sheets("교선-교직 (teaching.ts)", excel_elective_teach, src_teaching, "teaching.ts")

# 전공 비교 - 전필과 전선 분리
excel_major_req = [m for m in excel_major if m["category"] == "전필"]
excel_major_elec = [m for m in excel_major if m["category"] == "전선"]
excel_major_gyopil = [m for m in excel_major if m["category"] == "교필"]
excel_major_gyoseon = [m for m in excel_major if m["category"] == "교선"]
excel_major_semester = [m for m in excel_major if m["category"] == "학기"]

compare_sheets("전필 (major_required.ts)", excel_major_req, src_major_required, "major_required.ts")
compare_sheets("전선 (major_elective.ts)", excel_major_elec, src_major_elective, "major_elective.ts")

# 전공 시트 내 교필/교선/학기 카테고리
print(f"\n[참고] 전공 시트 내 타 카테고리 (core/electives/semester에서 관리):")
print(f"   - 교필: {len(excel_major_gyopil)}개 (core.ts와 중복)")
print(f"   - 교선: {len(excel_major_gyoseon)}개 (electives.ts와 중복)")
print(f"   - 학기: {len(excel_major_semester)}개 → semester.ts와 비교:")
compare_sheets("학기 (semester.ts)", excel_major_semester, src_semester, "semester.ts")

# 코드쉐어 비교
all_src = src_core + src_electives + src_major_required + src_major_elective + src_semester + src_normal_electives + src_teaching + src_online
compare_sheets("코드쉐어", excel_codeshare, all_src, "multiple")

# 마이크로디그리 비교
compare_sheets("마이크로디그리", excel_micro, all_src, "multiple")

# 추가 파일들
print(f"\n{'='*60}")
print("[추가 소스 파일 과목 수]")
print(f"{'='*60}")
print(f"  semester.ts: {len(src_semester)}개")
print(f"  normal_electives.ts: {len(src_normal_electives)}개")
print(f"  teaching.ts: {len(src_teaching)}개")
print(f"  online.ts: {len(src_online)}개")

# 전체 요약
print(f"\n{'='*60}")
print("[전체 요약]")
print(f"{'='*60}")
all_excel_ids = set()
for data in [excel_core, excel_elective, excel_major, excel_codeshare, excel_micro]:
    for e in data:
        all_excel_ids.add(e["id"])

all_src_ids = set()
for s in all_src:
    all_src_ids.add(normalize_id(s["id"]))

print(f"  엑셀 고유 과목(id) 수: {len(all_excel_ids)}개")
print(f"  소스 고유 과목(id) 수: {len(all_src_ids)}개")
print(f"  엑셀에만 있는 과목: {len(all_excel_ids - all_src_ids)}개")
print(f"  소스에만 있는 과목: {len(all_src_ids - all_excel_ids)}개")
print(f"  공통 과목: {len(all_excel_ids & all_src_ids)}개")
