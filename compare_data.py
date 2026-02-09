#!/usr/bin/env python3
"""Compare Excel source data with TypeScript course data files."""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import re
import os
import json
import openpyxl

EXCEL_PATH = r'26-1 수강편람 (4차).xlsx'
COURSES_DIR = r'src\data\courses'

def parse_ts_courses(filepath):
    """Extract course entries from a TypeScript file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    courses = []
    # Match each object block
    pattern = re.compile(r"\{[^{}]*?id:\s*'([^']+)'[^{}]*?\}", re.DOTALL)
    for m in pattern.finditer(content):
        block = m.group(0)
        course = {}
        # Extract key fields
        for key in ['id', 'code', 'section', 'name', 'category', 'timeRaw', 'roomRaw',
                     'creditDetail', 'college', 'department', 'major', 'year', 'note']:
            km = re.search(rf"{key}:\s*'([^']*)'", block)
            if km:
                course[key] = km.group(1)
        # Extract professors array
        pm = re.search(r"professors:\s*\[([^\]]*)\]", block)
        if pm:
            profs = re.findall(r"'([^']*)'", pm.group(1))
            course['professors'] = profs
        # Extract credits
        cm = re.search(r"credits:\s*(\d+)", block)
        if cm:
            course['credits'] = int(cm.group(1))
        # Extract isMicrodegree
        mm = re.search(r"isMicrodegree:\s*(true|false)", block)
        if mm:
            course['isMicrodegree'] = mm.group(1) == 'true'
        # Extract microdegreeNames
        mn = re.search(r"microdegreeNames:\s*\[([^\]]*)\]", block)
        if mn:
            names = re.findall(r"'([^']*)'", mn.group(1))
            course['microdegreeNames'] = names

        courses.append(course)
    return courses

def read_excel_data(wb):
    """Extract all course data from the Excel workbook."""
    all_courses = []

    # Sheet: 교필
    ws = wb['교필']
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        name, category, code, section, credit_detail = vals[0], vals[1], vals[2], vals[3], vals[4]
        college, dept, year, professor, time_raw, room_raw = vals[5], vals[6], vals[7], vals[8], vals[9], vals[10]
        if not name or not code:
            continue
        # Check if it's a sub-header row
        if name == '과목명':
            continue
        all_courses.append({
            'sheet': '교필',
            'id': f'{code}-{str(section).zfill(2)}',
            'code': str(code),
            'section': str(section).zfill(2),
            'name': str(name).strip(),
            'category': str(category).strip() if category else '',
            'creditDetail': str(credit_detail).strip() if credit_detail else '',
            'college': str(college).strip() if college else '',
            'department': str(dept).strip() if dept else '',
            'year': str(year).strip() if year else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
        })

    # Sheet: 교선
    ws = wb['교선']
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        name, category, code, section, credit_detail = vals[0], vals[1], vals[2], vals[3], vals[4]
        professor, time_raw, room_raw, note = vals[5], vals[6], vals[7], vals[8]
        if not name or not code:
            continue
        if name == '교과목명':
            continue
        all_courses.append({
            'sheet': '교선',
            'id': f'{code}-{str(section).zfill(2)}',
            'code': str(code),
            'section': str(section).zfill(2),
            'name': str(name).strip(),
            'category': str(category).strip() if category else '',
            'creditDetail': str(credit_detail).strip() if credit_detail else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
            'note': str(note).strip() if note else '',
        })

    # Sheet: 전공
    ws = wb['전공']
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        college, dept_division, major_name, year = vals[0], vals[1], vals[2], vals[3]
        code, section, name, category, credit_detail = vals[4], vals[5], vals[6], vals[7], vals[8]
        professor, time_raw, room_raw, note = vals[9], vals[10], vals[11], vals[12]
        if not name or not code:
            continue
        if name == '과목 명':
            continue
        all_courses.append({
            'sheet': '전공',
            'id': f'{code}-{str(section).zfill(2)}',
            'code': str(code),
            'section': str(section).zfill(2),
            'name': str(name).strip(),
            'category': str(category).strip() if category else '',
            'creditDetail': str(credit_detail).strip() if credit_detail else '',
            'college': str(college).strip() if college else '',
            'department': str(dept_division).strip() if dept_division else '',
            'major': str(major_name).strip() if major_name else '',
            'year': str(year).strip() if year else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
            'note': str(note).strip() if note else '',
        })

    # Sheet: 코드쉐어
    ws = wb['코드쉐어']
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        dept_name, major_field, major_dept = vals[0], vals[1], vals[2]
        code, section, credit_detail, year = vals[3], vals[4], vals[5], vals[6]
        category, professor, time_raw, room_raw = vals[7], vals[8], vals[9], vals[10]
        if not code:
            continue
        if str(code).strip() == '학수번호':
            continue
        all_courses.append({
            'sheet': '코드쉐어',
            'id': f'{code}-{str(section).zfill(2)}',
            'code': str(code),
            'section': str(section).zfill(2),
            'department': str(dept_name).strip() if dept_name else '',
            'major': str(major_dept).strip() if major_dept else '',
            'creditDetail': str(credit_detail).strip() if credit_detail else '',
            'year': str(year).strip() if year else '',
            'category': str(category).strip() if category else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
        })

    # Sheet: 마이크로디그리
    ws = wb['마이크로디그리']
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        micro_name, code, dept_name, section = vals[0], vals[1], vals[2], vals[3]
        credit_detail, year, category, professor = vals[4], vals[5], vals[6], vals[7]
        time_raw, room_raw, note = vals[8], vals[9], vals[10]
        if not code:
            continue
        if str(code).strip() == '학수번호':
            continue
        all_courses.append({
            'sheet': '마이크로디그리',
            'id': f'{code}-{str(section).zfill(2)}',
            'code': str(code),
            'section': str(section).zfill(2),
            'microdegreeName': str(micro_name).strip() if micro_name else '',
            'department': str(dept_name).strip() if dept_name else '',
            'creditDetail': str(credit_detail).strip() if credit_detail else '',
            'year': str(year).strip() if year else '',
            'category': str(category).strip() if category else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
            'note': str(note).strip() if note else '',
        })

    return all_courses

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    excel_courses = read_excel_data(wb)

    print(f"=== Excel Data Summary ===")
    by_sheet = {}
    for c in excel_courses:
        by_sheet[c['sheet']] = by_sheet.get(c['sheet'], 0) + 1
    for sheet, count in by_sheet.items():
        print(f"  {sheet}: {count} entries")
    print(f"  Total: {len(excel_courses)} entries")

    # Build Excel lookup by id
    excel_by_id = {}
    for c in excel_courses:
        excel_by_id[c['id']] = c

    # Parse TypeScript files
    ts_files = ['core.ts', 'electives.ts', 'major_required.ts', 'major_elective.ts',
                'semester.ts', 'normal_electives.ts', 'teaching.ts', 'online.ts']
    all_ts_courses = {}
    ts_by_file = {}

    print(f"\n=== TypeScript Data Summary ===")
    for f in ts_files:
        path = os.path.join(COURSES_DIR, f)
        courses = parse_ts_courses(path)
        ts_by_file[f] = courses
        for c in courses:
            all_ts_courses[c.get('id', '')] = (f, c)
        print(f"  {f}: {len(courses)} entries")
    print(f"  Total (before dedup): {len(all_ts_courses)} unique IDs")

    # Excel unique IDs
    excel_ids = set(excel_by_id.keys())
    ts_ids = set(all_ts_courses.keys())

    # Find missing (in Excel but not in TS)
    missing_in_ts = excel_ids - ts_ids
    extra_in_ts = ts_ids - excel_ids
    common = excel_ids & ts_ids

    print(f"\n=== Comparison ===")
    print(f"  Excel unique IDs: {len(excel_ids)}")
    print(f"  TS unique IDs: {len(ts_ids)}")
    print(f"  Common: {len(common)}")
    print(f"  In Excel but NOT in TS: {len(missing_in_ts)}")
    print(f"  In TS but NOT in Excel: {len(extra_in_ts)}")

    if missing_in_ts:
        print(f"\n=== MISSING IN TS (in Excel but not in course files) ===")
        # Group by sheet
        by_sheet_missing = {}
        for mid in sorted(missing_in_ts):
            c = excel_by_id[mid]
            sheet = c['sheet']
            if sheet not in by_sheet_missing:
                by_sheet_missing[sheet] = []
            by_sheet_missing[sheet].append(c)
        for sheet, courses in by_sheet_missing.items():
            print(f"\n  --- {sheet} ({len(courses)} missing) ---")
            for c in courses[:50]:  # limit output
                print(f"    {c['id']}: {c.get('name', '?')} | {c.get('category', '')} | {c.get('professor', '')} | {c.get('timeRaw', '')} | {c.get('roomRaw', '')}")
            if len(courses) > 50:
                print(f"    ... and {len(courses)-50} more")

    if extra_in_ts:
        print(f"\n=== EXTRA IN TS (in TS but not in Excel) ===")
        for eid in sorted(extra_in_ts)[:50]:
            tsf, c = all_ts_courses[eid]
            print(f"  {eid}: {c.get('name', '?')} (file: {tsf})")
        if len(extra_in_ts) > 50:
            print(f"  ... and {len(extra_in_ts)-50} more")

    # Compare fields for common entries
    print(f"\n=== FIELD DIFFERENCES (common entries) ===")
    diffs_count = 0
    for cid in sorted(common):
        ec = excel_by_id[cid]
        tsf, tc = all_ts_courses[cid]
        diffs = []

        # Compare name
        if ec.get('name', '') != tc.get('name', ''):
            diffs.append(f"name: Excel='{ec.get('name')}' vs TS='{tc.get('name')}'")

        # Compare category
        if ec.get('category', '') != tc.get('category', ''):
            diffs.append(f"category: Excel='{ec.get('category')}' vs TS='{tc.get('category')}'")

        # Compare creditDetail
        if ec.get('creditDetail', '') != tc.get('creditDetail', ''):
            diffs.append(f"creditDetail: Excel='{ec.get('creditDetail')}' vs TS='{tc.get('creditDetail')}'")

        # Compare professor
        excel_prof = ec.get('professor', '')
        ts_profs = tc.get('professors', [])
        ts_prof_str = ','.join(ts_profs)
        if excel_prof != ts_prof_str:
            diffs.append(f"professor: Excel='{excel_prof}' vs TS='{ts_prof_str}'")

        # Compare timeRaw
        if ec.get('timeRaw', '') != tc.get('timeRaw', ''):
            diffs.append(f"timeRaw: Excel='{ec.get('timeRaw')}' vs TS='{tc.get('timeRaw')}'")

        # Compare roomRaw
        if ec.get('roomRaw', '') != tc.get('roomRaw', ''):
            diffs.append(f"roomRaw: Excel='{ec.get('roomRaw')}' vs TS='{tc.get('roomRaw')}'")

        if diffs:
            diffs_count += 1
            if diffs_count <= 100:
                print(f"\n  {cid} ({ec.get('name', '?')}) [file: {tsf}]:")
                for d in diffs:
                    print(f"    {d}")

    if diffs_count > 100:
        print(f"\n  ... and {diffs_count - 100} more entries with differences")
    print(f"\n  Total entries with differences: {diffs_count} out of {len(common)} common entries")

    # Verify 전공 sheet category distribution in TS
    print(f"\n=== 전공 Sheet Category Distribution ===")
    major_cats_excel = {}
    for c in excel_courses:
        if c['sheet'] == '전공':
            cat = c.get('category', '')
            major_cats_excel[cat] = major_cats_excel.get(cat, 0) + 1
    print(f"  Excel: {major_cats_excel}")

    # Check TS category mapping
    ts_cats = {}
    for f in ts_files:
        for c in ts_by_file[f]:
            cat = c.get('category', '')
            if f not in ts_cats:
                ts_cats[f] = {}
            ts_cats[f][cat] = ts_cats[f].get(cat, 0) + 1
    for f, cats in ts_cats.items():
        print(f"  {f}: {cats}")

if __name__ == '__main__':
    main()
