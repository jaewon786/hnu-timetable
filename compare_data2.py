#!/usr/bin/env python3
"""Compare Excel source data with TypeScript course data files - v2."""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import re, os
import openpyxl

EXCEL_PATH = r'26-1 수강편람 (4차).xlsx'
COURSES_DIR = r'src\data\courses'

def normalize_id(raw_id):
    """Normalize ID to code-section format without zero padding."""
    parts = raw_id.split('-')
    if len(parts) == 2:
        code, section = parts
        # Remove leading zeros from section
        return f"{code}-{section.lstrip('0') or '0'}"
    return raw_id

def parse_ts_ids_and_data(filepath):
    """Extract id and basic info from TS file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    entries = {}
    # Find all id values
    ids = re.findall(r"id:\s*'([^']+)'", content)
    # For each id, try to find name too
    # Use a more robust approach: find blocks starting with { containing id:
    # Split by '  {' or '\n  {' patterns to find entries
    # Actually, just extract id->name pairs
    id_name_pairs = re.findall(r"id:\s*'([^']+)'[^}]*?name:\s*'([^']*)'", content, re.DOTALL)

    for raw_id, name in id_name_pairs:
        nid = normalize_id(raw_id)
        entries[nid] = {'id': raw_id, 'name': name}

    # Some IDs might not have matched with name (due to nested braces)
    for raw_id in ids:
        nid = normalize_id(raw_id)
        if nid not in entries:
            entries[nid] = {'id': raw_id, 'name': '?'}

    return entries

def read_excel_ids(wb):
    """Extract all course IDs from Excel."""
    all_courses = {}

    # Sheet: 교필
    ws = wb['교필']
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        category = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        section = ws.cell(r, 4).value
        credit = ws.cell(r, 5).value
        professor = ws.cell(r, 9).value
        time_raw = ws.cell(r, 10).value
        room_raw = ws.cell(r, 11).value
        if not name or not code or name == '과목명':
            continue
        nid = f"{code}-{str(section).lstrip('0') or '0'}"
        all_courses[nid] = {
            'sheet': '교필', 'name': str(name).strip(), 'category': str(category).strip() if category else '',
            'creditDetail': str(credit).strip() if credit else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
        }

    # Sheet: 교선
    ws = wb['교선']
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        category = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        section = ws.cell(r, 4).value
        credit = ws.cell(r, 5).value
        professor = ws.cell(r, 6).value
        time_raw = ws.cell(r, 7).value
        room_raw = ws.cell(r, 8).value
        note = ws.cell(r, 9).value
        if not name or not code or name == '교과목명':
            continue
        nid = f"{code}-{str(section).lstrip('0') or '0'}"
        all_courses[nid] = {
            'sheet': '교선', 'name': str(name).strip(), 'category': str(category).strip() if category else '',
            'creditDetail': str(credit).strip() if credit else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
            'note': str(note).strip() if note else '',
        }

    # Sheet: 전공
    ws = wb['전공']
    for r in range(3, ws.max_row + 1):
        college = ws.cell(r, 1).value
        dept = ws.cell(r, 2).value
        major = ws.cell(r, 3).value
        year = ws.cell(r, 4).value
        code = ws.cell(r, 5).value
        section = ws.cell(r, 6).value
        name = ws.cell(r, 7).value
        category = ws.cell(r, 8).value
        credit = ws.cell(r, 9).value
        professor = ws.cell(r, 10).value
        time_raw = ws.cell(r, 11).value
        room_raw = ws.cell(r, 12).value
        note = ws.cell(r, 13).value
        if not name or not code or name == '과목 명':
            continue
        nid = f"{code}-{str(section).lstrip('0') or '0'}"
        all_courses[nid] = {
            'sheet': '전공', 'name': str(name).strip(), 'category': str(category).strip() if category else '',
            'creditDetail': str(credit).strip() if credit else '',
            'college': str(college).strip() if college else '',
            'department': str(dept).strip() if dept else '',
            'major': str(major).strip() if major else '',
            'year': str(year).strip() if year else '',
            'professor': str(professor).strip() if professor else '',
            'timeRaw': str(time_raw).strip() if time_raw else '',
            'roomRaw': str(room_raw).strip() if room_raw else '',
            'note': str(note).strip() if note else '',
        }

    # Sheet: 코드쉐어
    ws = wb['코드쉐어']
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 4).value
        section = ws.cell(r, 5).value
        category = ws.cell(r, 8).value
        professor = ws.cell(r, 9).value
        time_raw = ws.cell(r, 10).value
        room_raw = ws.cell(r, 11).value
        if not code or str(code).strip() == '학수번호':
            continue
        nid = f"{code}-{str(section).lstrip('0') or '0'}"
        # Don't overwrite if already exists from 전공 sheet
        if nid not in all_courses:
            all_courses[nid] = {
                'sheet': '코드쉐어',
                'name': '(코드쉐어)',
                'category': str(category).strip() if category else '',
                'professor': str(professor).strip() if professor else '',
                'timeRaw': str(time_raw).strip() if time_raw else '',
                'roomRaw': str(room_raw).strip() if room_raw else '',
            }

    # Sheet: 마이크로디그리 - these are references, not separate courses
    # They point to courses that should already exist in other sheets

    return all_courses

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    excel_courses = read_excel_ids(wb)

    # Count by sheet
    sheet_counts = {}
    for c in excel_courses.values():
        s = c['sheet']
        sheet_counts[s] = sheet_counts.get(s, 0) + 1
    print("=== Excel Unique IDs by Sheet ===")
    for s, cnt in sheet_counts.items():
        print(f"  {s}: {cnt}")
    print(f"  Total unique: {len(excel_courses)}")

    # Parse TS files
    ts_files = ['core.ts', 'electives.ts', 'major_required.ts', 'major_elective.ts',
                'semester.ts', 'normal_electives.ts', 'teaching.ts', 'online.ts']

    all_ts = {}
    print("\n=== TypeScript Entries ===")
    for f in ts_files:
        path = os.path.join(COURSES_DIR, f)
        entries = parse_ts_ids_and_data(path)
        for nid, data in entries.items():
            if nid not in all_ts:
                all_ts[nid] = (f, data)
        print(f"  {f}: {len(entries)} entries")
    print(f"  Total unique: {len(all_ts)}")

    # Compare
    excel_ids = set(excel_courses.keys())
    ts_ids = set(all_ts.keys())
    missing_in_ts = excel_ids - ts_ids
    extra_in_ts = ts_ids - excel_ids
    common = excel_ids & ts_ids

    print(f"\n=== Comparison (normalized IDs) ===")
    print(f"  Excel unique: {len(excel_ids)}")
    print(f"  TS unique: {len(ts_ids)}")
    print(f"  Common: {len(common)}")
    print(f"  Missing in TS: {len(missing_in_ts)}")
    print(f"  Extra in TS: {len(extra_in_ts)}")

    # Missing in TS - grouped by sheet and category
    print(f"\n=== MISSING IN TS (grouped by sheet+category) ===")
    missing_by_group = {}
    for mid in missing_in_ts:
        c = excel_courses[mid]
        key = f"{c['sheet']} / {c.get('category', '?')}"
        if key not in missing_by_group:
            missing_by_group[key] = []
        missing_by_group[key].append((mid, c))
    for group in sorted(missing_by_group.keys()):
        items = missing_by_group[group]
        print(f"\n  [{group}] - {len(items)} missing")
        for mid, c in sorted(items)[:10]:
            print(f"    {mid}: {c.get('name', '?')} | {c.get('professor', '')} | time={c.get('timeRaw', '')} | room={c.get('roomRaw', '')}")
        if len(items) > 10:
            print(f"    ... and {len(items)-10} more")

    # Extra in TS
    print(f"\n=== EXTRA IN TS (not in Excel) ===")
    for eid in sorted(extra_in_ts):
        tsf, data = all_ts[eid]
        print(f"  {eid}: {data.get('name', '?')} (file: {tsf})")

    # Field differences for common entries
    print(f"\n=== FIELD DIFFERENCES (sample) ===")
    diff_count = 0
    for cid in sorted(common):
        ec = excel_courses[cid]
        tsf, tc = all_ts[cid]
        diffs = []

        # Compare name
        ts_name = tc.get('name', '')
        excel_name = ec.get('name', '')
        if excel_name and ts_name and excel_name != ts_name:
            # Ignore Ⅰ vs I differences
            if excel_name.replace('Ⅰ', 'I').replace('Ⅱ', 'II').replace('Ⅲ', 'III') != ts_name.replace('Ⅰ', 'I').replace('Ⅱ', 'II').replace('Ⅲ', 'III'):
                diffs.append(f"name: Excel='{excel_name}' vs TS='{ts_name}'")

        if diffs:
            diff_count += 1
            if diff_count <= 30:
                print(f"  {cid} [{tsf}]: {'; '.join(diffs)}")

    print(f"\n  Total with meaningful name differences: {diff_count}")

if __name__ == '__main__':
    main()
